from __future__ import annotations

import os
from pathlib import Path
from typing import Dict
import logging

from openpyxl import load_workbook
from dotenv import load_dotenv

from .helpers.cdisc_connector import CDISCCTConnector
from .helpers.evs_connector import NCIEVSConnector
from .models.eap import EnumeratedValue, Document, Enumeration

load_dotenv()

from .loader import load_expanded_dir
from .models.sqlite_loader import load_from_file
from .models.usdm_ct import CodeList, DDFEntity, PermissibleValue

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def load_usdm_ct(filename: str):
    """
    The CT contains extra descriptive metadata for the Entities and Attributes
    * it also contains the value sets
    * plus references to external value sets
    """
    ct = CDISCCTConnector(os.environ["CDISC_LIBRARY_API_TOKEN"])
    # If required
    # evs = NCIEVSConnector()

    print("Loading USDM CT")
    if not os.path.exists(filename):
        raise FileNotFoundError(f"File not found: {filename}")
    wbk = load_workbook(filename, read_only=True)
    # load the entities
    _ent_attr = wbk["DDF Entities&Attributes"]
    entities = {}
    headers = []
    for idx, row in enumerate(_ent_attr.iter_rows(min_row=1, values_only=True)):
        if idx == 0:
            for cell in row:
                headers.append(cell)
            continue
        # map the row to a dictionary
        _row = dict(zip(headers, row))
        _name = _row["Entity Name"]
        _role = _row["Role"]
        _entity = DDFEntity.from_dict(_row)
        if _entity.role in ("Entity"):
            if not _name in entities:
                # add an id attribute
                _id = DDFEntity.from_dict(
                    {
                        "Entity Name": f"{_name}",
                        "Role": "Attribute",
                        "Logical Data Model Name": "id",
                        "Definition": "One or more characters used to identify, name, or characterize the nature, "
                        "properties, or contents of a thing.",
                        "CT Item Preferred Name": "id",
                        "Has Value List": "N",
                        "NCI C-code": "C25364",
                        "Inherited From": "",
                        "Synonym(s)": "Identifier; Unique Identifier",
                    }
                )
                _entity.attributes.append(_id)
                entities[_name] = _entity

        elif _entity.role in ("Relationship"):
            entities[_name].relationships.append(_entity)
        elif _entity.role in ("Complex Datatype Relationship"):
            entities[_name].complex_datatype_relationships.append(_entity)
        elif _entity.role in ("Attribute"):
            entities[_name].attributes.append(_entity)
        else:
            raise ValueError(f"Unknown entity type: {_role}")
    print("Loading USDM CT Value Sets")
    # load the values from the values sheet
    _value_sets = wbk["DDF valid value sets"]
    codelists = {}
    for row in _value_sets.iter_rows(min_row=7, values_only=True):
        # get the entity
        codeset = PermissibleValue.from_row(row)
        if codeset.codelist_c_code not in codelists:
            # create a new codelist
            codelists[codeset.codelist_c_code] = CodeList.from_pvalue(codeset)
        codelists[codeset.codelist_c_code].add_item(codeset)

    for entity in entities.values():  # type: DDFEntity
        for attr in entity.all_attributes.values():
            if attr.has_value_list:
                if attr.external_code_list is not None:
                    if attr.external_code_list.startswith("C"):
                        if attr.external_code_list == "CNEW":
                            # create a new codelist placeholder
                            codelist = CodeList(
                                entity_name=entity.entity_name,
                                attribute_name=attr.logical_data_model_name,
                                concept_c_code=f"CNEW-{attr.logical_data_model_name}",
                                submission_value=f"CNEW ({entity.entity_name}.{attr.logical_data_model_name})",
                                alternate_name=f"New Code for {entity.entity_name}.{attr.logical_data_model_name}",
                                extensible=True,
                            )
                            codelists[f"CNEW-{attr.logical_data_model_name}"] = codelist
                            attr.codelist = codelist
                        elif attr.external_code_list not in codelists:
                            # pull the codelist from the API
                            print(
                                f"Retrieving codelist {attr.external_code_list} for {entity.entity_name}.{attr.logical_data_model_name}"
                            )
                            codelist = ct.retrieve_valueset(attr.external_code_list)
                            if codelist:
                                codelists[attr.external_code_list] = codelist
                            else:
                                print(
                                    f"Retrieving codelist {attr.external_code_list} for {entity.entity_name}.{attr.logical_data_model_name} failed"
                                )
                            # bind the codelist to the attribute
                            attr.codelist = codelist
                        else:
                            # bind the codelist to the attribute
                            attr.codelist = codelists[attr.external_code_list]
                else:
                    print(
                        f"Extracting {attr.value_list}  for {attr.logical_data_model_name} failed"
                    )

    return entities, codelists


def main_usdm(
    source_dir_or_file: str, controlled_term: str, output_dir: str, gen: Dict[str, bool]
):
    NAMESPACE = "https://cdisc.org/usdm"
    if Path(source_dir_or_file).is_file():
        assert Path(source_dir_or_file).suffix == ".qea", "Only QEA files are supported"
        logger.info(f"Loading from file {source_dir_or_file}")
        # handle the cases, release_3-5-0 and v3.5.0
        version = Path(source_dir_or_file).stem[:-9]        
        name = "USDM" + "_" + version
        document = load_from_file(source_dir_or_file, prefix=NAMESPACE, name=name)
    else:
        logger.info(f"Loading from directory {source_dir_or_file}")
        version = Path(source_dir_or_file).name.split("_")[0]
        name = (
            os.path.basename(os.path.dirname(source_dir_or_file))
            if source_dir_or_file.endswith("/")
            else os.path.basename(source_dir_or_file)
        )
        document = load_expanded_dir(source_dir_or_file)
    document.version = version
    document.root_item = "Study"
    document.description = (
        "Unified Study Definitions Model version 3. The industry standard "
        "for digital protocol content, defining reusable protocols expressing scientific questions "
        "via structure and concepts that are reused and referenced downstream, creating a knowledge graph."
    )
    document.add_prefix("usdm", NAMESPACE)
    document.add_prefix("ncit", "https://ncicb.nci.nih.gov/xml/owl/EVS/Thesaurus.owl")
    # loaded content from the USDM CT
    ct_content, codelists = load_usdm_ct(controlled_term)
    # update the document with the CT content
    for entity in document.objects:
        if entity.object_type == "Class":
            # add the codelist to the class
            if entity.name in ct_content:
                # print(f"Adding CT content for {object.name}")
                _content = ct_content[entity.name]  # type: DDFEntity
                if _content.definition:
                    # add the definition
                    entity.definition = _content.definition
                if _content.nci_c_code:
                    # add the nci code
                    entity.reference_url = _content.nci_c_code
                if _content.preferred_term:
                    entity.preferred_term = _content.preferred_term
                if _content.synonyms:
                    entity.synonyms = _content.synonyms
                # combine the attributes
                for attr in entity.all_attributes:
                    if _content.get_attribute(attr.name):
                        # match the attribute by name
                        # print(f"Adding CT content for {object.name}.{attr.name}")
                        _attr = _content.get_attribute(attr.name)
                        if _attr.definition:
                            attr.definition = _attr.definition
                        if _attr.nci_c_code:
                            attr.reference_url = _attr.nci_c_code
                        if _attr.codelist:
                            print(
                                f"Adding codelist {_attr.codelist.concept_c_code} to {entity.name}.{attr.name}"
                            )
                            attr.codelist = _attr.codelist
                        if _attr.synonyms:
                            attr.synonyms = _attr.synonyms
                        if _attr.preferred_term:
                            attr.preferred_term = _attr.preferred_term
        else:
            print(f"Skipping {entity.object_type} {entity.name}")
    # for concept in ct_content.values():
    #     if concept.definition:
    #         definitions[concept.entity_name] = concept.definition
    #     for attr in concept.attributes:
    #         if attr.definition:
    #             definitions[attr.logical_data_model_name] = attr.definition
    for aspect, genflag in gen.items():
        print("Checking generation of ", aspect, "as", genflag)
        if genflag:
            if aspect == "prisma":
                from .render.render_prisma import generate as generate_prisma

                generate_prisma(
                    document.name, document, ct_content, codelists, output_dir
                )
            elif aspect == "linkml":
                from .render.render_linkml import (
                    generate_schema_builder as generate_linkml,
                )

                generate_linkml(
                    name=document.name,
                    document=document,
                    schema_id="https://cdisc.org/usdm/" + document.name,
                    output_dir=output_dir,
                )
            elif aspect == "shapes":
                from .render.render_shapes import generate as generate_shapes

                generate_shapes(
                    document.name, document, ct_content, codelists, output_dir
                )
            else:
                raise ValueError(f"Unknown aspect: {aspect}")
    # always generate the workbook
    from .render.render_usdm_workbook import generate as generate_workbook

    _ = generate_workbook(document.name, document, ct_content, codelists, output_dir)
