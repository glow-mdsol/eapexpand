from __future__ import annotations

from dataclasses import dataclass, field
import os

from openpyxl import Workbook
from datetime import datetime
from typing import Dict, Optional, List

from eapexpand.loader import load_expanded_dir


def generate(
    name: str,
    packages: dict,
    objects: dict,
    attributes: dict,
    connectors: dict,
    output_dir: Optional[str] = "output",
):
    """
    Generates the Excel Representation of the model
    :param name: The name of the model - guides what the output file is called
    """
    # subset by packages
    _partitions = {}
    for _object in objects.values():
        if _object.package_id in packages:
            _package = packages.get(_object.package_id)
            _partitions.setdefault(_package.name, []).append(_object)

    doc = Workbook()
    
    for _partition_id, objects in _partitions.items():
        _package = packages.get(_partition_id)
    sheet = doc.active
    sheet.title = "Objects"
    for idx, column in enumerate(
        ("Class", "Attribute", "Type", "Cardinality", "Class Note")
    ):
        sheet.cell(row=1, column=idx + 1).value = column
    row_num = 2
    for obj in objects.values():
        _output = {}
        if obj.object_type == "Class":
            _object_id = obj.object_id
            _attributes = sorted(
                [
                    attributes[attr_id]
                    for attr_id in attributes
                    if attributes[attr_id].object_id == _object_id
                ]
            )
            # write the entity
            sheet.cell(row=row_num, column=1).value = obj.name
            if obj.note:
                sheet.cell(row=row_num, column=5).value = str(obj.note)
            row_num += 1
            for _attribute in _attributes:
                attrib = _output.setdefault(_attribute.name, {})
                if not attrib:
                    attrib = dict(
                        attribute_name=_attribute.name,
                        attribute_type=_attribute.attribute_type,
                        attribute_cardinality=_attribute.cardinality
                    )
                # TODO - upsert
                _output[_attribute.name] = attrib
                # _output[_attribute.name] = dict(name)
                #     sheet.cell(row=row_num, column=1).value = obj.name
                #     sheet.cell(row=row_num, column=2).value = _attribute.name
                #     sheet.cell(row=row_num, column=3).value = _attribute.attribute_type
                #     sheet.cell(row=row_num, column=4).value = "*" if "List" in _attribute.attribute_type else ""
                #     if _init:
                #        sheet.cell(row=row_num, column=5).value = obj.note if obj.note else ""
                #        _init = False
                #     row_num += 1
            connections = [
                connectors[cid]
                for cid in connectors
                if connectors[cid].start_object_id == _object_id
            ]
            # for connection in connections:
            #     if connection.connector_type not in ("Association", "Generalization"):
            #         continue
            #     attrib = _output.setdefault(connection.name, {})
            #     if not attrib:
            #         attrib = dict(
            #             attribute_name=connection.name,
            #             attribute_type=objects[connection.end_object_id].name,
            #             attribute_cardinality=connection.source_card,
            #             note=None,
            #         )
            #     _output[connection.name] = attrib
                # _attributes = sorted([attributes[aid] for aid in attributes if attributes[aid].object_id == _connector.end_object_id])
                # for _attribute in _attributes:
                #     attrib = _output.setdefault(_attribute.name, {})
                #     sheet.cell(row=row_num, column=1).value = obj.name
                #     sheet.cell(row=row_num, column=2).value = _attribute.name
                #     sheet.cell(row=row_num, column=3).value = _attribute.attribute_type
                #     sheet.cell(row=row_num, column=4).value = _connector.dest_card if _connector.dest_card else ""
                # sheet.cell(row=row_num, column=5).value = _attribute.note if _attribute.note else ""
            for offset, attrib in enumerate(_output.values()):
                sheet.cell(row=row_num, column=1).value = obj.name
                sheet.cell(row=row_num, column=2).value = attrib["attribute_name"]
                sheet.cell(row=row_num, column=3).value = attrib["attribute_type"]
                sheet.cell(row=row_num, column=4).value = attrib[
                    "attribute_cardinality"
                ]
                row_num += 1
    # create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    fname = os.path.join(output_dir, f"{name}.xlsx")
    doc.save(fname)
    print(f"Generated Excel file: {fname}")


def main(source_dir: str, output_dir: str):
    """
    Main entry point
    """
    name = (
        os.path.basename(os.path.dirname(source_dir))
        if source_dir.endswith("/")
        else os.path.basename(source_dir)
    )
    objects, attributes, connectors = load_expanded_dir(source_dir)
    generate(name, objects, attributes, connectors, output_dir=output_dir)
