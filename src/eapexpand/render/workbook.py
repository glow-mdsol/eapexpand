import os
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def generate(
    name: str,
    objects: dict,
    output_dir: Optional[str] = "output",
):
    """
    Generates the Excel Representation of the model
    :param name: The name of the model - guides what the output file is called
    """
    HEADERS =  ("Package", "Class", "Attribute", "Type", "Cardinality", "Class Note")
    packages = {x.package_id: x for x in objects.values() if x.object_type == "Package"}
    # subset by packages
    _partitions = {}
    for _object in objects.values():
        if _object.package_id in packages:
            _package = packages.get(_object.package_id)
            _partitions.setdefault(_package.name, []).append(_object)
        else:
            pass
            # print("Orphaned Object: ", _object.name, " -> ", _object.package_id)

    doc = Workbook()
    # have the specs for the columns internally
    cols = {}
    def write_cell(worksheet, row, column, value, header = False):
        columns = cols.setdefault(worksheet.title, {})
        columns[column] = max([len(value), columns.get(column, 10)])
        cols[worksheet.title] = columns
        worksheet.cell(row, column).value = value
        worksheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical='top')
        if header:
            worksheet.cell(row, column).font = Font(bold=True)
    # write the packages
    wkst = doc.active
    wkst.title = "Packages"
    write_cell(wkst, 1, 1, "Package", True)
    write_cell(wkst, 1, 3, "Note", True)
    write_cell(wkst, 1, 2, "Parent", True)
    row_num = 2
    for _package in packages.values():
        write_cell(wkst, row_num, 1, _package.name)
        write_cell(wkst, row_num, 3, _package.note)
        if _package.parent:
            write_cell(wkst, row_num, 2, _package.parent.name)
        row_num += 1
    for _package_name, objects in _partitions.items():
        # Skip packages with only one object
        if len(objects) == 1:
            continue
        _sheet_name = _package_name[:30]
        # Limit on the Sheet Name length
        sheet = doc.create_sheet(_sheet_name)
        # write the header
        for idx, column in enumerate(
            HEADERS
        ):
            write_cell(sheet, 1, idx + 1, column, header=True)
        row_num = 2
        # iterate over the objects in the package
        for obj in objects:
            _output = {}
            if obj.object_type == "Class":
                # write the entity
                write_cell(sheet, row=row_num, column=1, value = _package_name)
                write_cell(sheet, row=row_num, column=2, value = obj.name)
                if obj.note:
                    write_cell(sheet, row=row_num, column=6, value = str(obj.note))
                row_num += 1
                for _attribute in obj.attributes:
                    attrib = _output.setdefault(_attribute.name, {})
                    if not attrib:
                        attrib = dict(
                            attribute_name=_attribute.name,
                            attribute_type=_attribute.attribute_type,
                            attribute_cardinality=_attribute.cardinality,
                            attribute_note=_attribute.notes,
                        )
                    # TODO - upsert
                    _output[_attribute.name] = attrib
                for _, attrib in enumerate(_output.values()):
                    write_cell(sheet, row=row_num, column=1, value = _package_name)
                    write_cell(sheet, row=row_num, column=2, value = obj.name)
                    write_cell(sheet, row=row_num, column=3, value = attrib["attribute_name"])
                    write_cell(sheet, row=row_num, column=4, value = attrib["attribute_type"])
                    write_cell(sheet, row=row_num, column=5, value = attrib[
                        "attribute_cardinality"
                    ])
                    write_cell(sheet, row=row_num, column=6, value = attrib[
                        "attribute_note"
                    ])
                    row_num += 1
            else:
                print("Skipping object: ", obj.name, " of type ", obj.object_type, " in package ", _package_name)
    # set the column widths
    for wks in doc.worksheets:
        _columns = cols.get(wks.title, {})
        for colnum, length in _columns.items():
            column_width = min([length, 50])
            wks.column_dimensions[get_column_letter(colnum)].width = column_width

    # create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    fname = os.path.join(output_dir, f"{name}.xlsx")
    doc.save(fname)
    print(f"Generated Excel file: {fname}")
