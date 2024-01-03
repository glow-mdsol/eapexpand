import os
from typing import Dict, Optional
from eapexpand.models.eap import Attribute, Connector, Document, Object

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADERS = [
    "Class",
    "Attribute",
    "Type",
    "Cardinality",
    "Class Note",
    "NCI C-code",
    "Preferred term",
    "Definition",
    "Codelist",
    "External Codelist",
]


def generate(
    name: str,
    document: Document,
    ct_content: dict,
    codelists: dict,
    output_dir: Optional[str] = "output",
):
    """
    Generates the Excel Representation of the model
    :param name: The name of the model - guides what the output file is called
    """
    print("Generating USDM Excel file")
    packages = {x.package_id: x for x in document.objects if x.object_type == "Package"}
    # remap the code attribute model to the code list model
    # subset by packages
    _partitions = {}
    for _object in document.objects:
        if _object.package_id in packages:
            _package = packages.get(_object.package_id)
            _partitions.setdefault(_package.name, []).append(_object)
        else:
            print("Orphaned Object: ", _object.name, " -> ", _object.package_id)
    cols = {}
    doc = Workbook()

    def write_cell(worksheet, row, column, value, header=False):
        _value = value if value else ""
        columns = cols.setdefault(worksheet.title, {})
        columns[column] = max([len(_value), columns.get(column, 10)])
        cols[worksheet.title] = columns
        worksheet.cell(row, column).value = _value
        worksheet.cell(row, column).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        if header:
            worksheet.cell(row, column).font = Font(bold=True)

    # Write the package sheet
    wkst = doc.active
    wkst.title = "Packages"
    write_cell(wkst, 1, 1, "Package", True)
    write_cell(wkst, 1, 2, "Parent", True)
    write_cell(wkst, 1, 3, "Note", True)
    row_num = 2
    for _package in packages.values():
        write_cell(wkst, row_num, 1, _package.name)
        write_cell(wkst, row_num, 3, _package.note)
        if _package.parent:
            write_cell(wkst, row_num, 2, _package.parent.name)
        row_num += 1
    for _package_name, pobjects in _partitions.items():
        if len(pobjects) == 1:
            continue
        _sheet_name = _package_name[:30]
        # Limit on the Sheet Name length
        sheet = doc.create_sheet(_sheet_name)
        # write the header
        for idx, column in enumerate(HEADERS):
            write_cell(sheet, 1, idx + 1, column, header=True)
        row_num = 2

        for obj in pobjects:  # type: Object
            _output = {}
            if obj.object_type == "Class":
                _ref = ct_content.get(obj.name)  # type: Entity
                if obj.generalizations:
                    _generalization = document.get_object(
                        obj.generalizations[0].end_object_id
                    )
                    _ref = ct_content.get(_generalization.name)
                else:
                    _generalization = None
                if _ref is None:
                    print("WARNING: Reference not found for " + obj.name)
                # _attributes = sorted(
                #     [
                #         attributes[attr_id]
                #         for attr_id in attributes
                #         if attributes[attr_id].object_id == _object_id
                #     ]
                # )
                # write the entity
                write_cell(sheet, row=row_num, column=1, value=obj.name)
                if _ref:
                    if _ref.definition:
                        write_cell(sheet, row=row_num, column=8, value=_ref.definition)
                    if _ref.nci_c_code:
                        write_cell(sheet, row=row_num, column=6, value=_ref.nci_c_code)
                    if _ref.preferred_term:
                        write_cell(
                            sheet, row=row_num, column=7, value=_ref.preferred_term
                        )
                else:
                    print(f"Reference not found for {obj.name}")
                if obj.note:
                    write_cell(sheet, row=row_num, column=5, value=str(obj.note))
                row_num += 1

                for _attribute in obj.attributes:  # type: Attribute
                    attrib = _output.setdefault(_attribute.name, {})
                    if not attrib:
                        if (
                            _generalization
                            and _attribute in _generalization.object_attributes
                        ):
                            _name = "* " + _attribute.name
                        else:
                            _name = _attribute.name
                        attrib = dict(
                            attribute_name=_name,
                            attribute_type=_attribute.attribute_type,
                            attribute_cardinality=_attribute.cardinality,
                            attribute_note=_attribute.note,
                        )
                        _attr_ref = _ref.get_attribute(_attribute.name)
                        if _attr_ref:
                            attrib["definition"] = _attr_ref.definition
                            attrib["c_code"] = _attr_ref.nci_c_code
                            attrib["pref_term"] = _attr_ref.preferred_term
                            if _attr_ref.has_value_list:
                                if _attr_ref.external_code_list:
                                    attrib[
                                        "external_value_list"
                                    ] = _attr_ref.external_code_list
                                elif _attr_ref.codelist_code:
                                    attrib["codelist"] = _attr_ref.codelist_code
                        else:      
                            print(
                                f"Reference not found for attribute {_attribute.name} in {_ref.entity_name}"
                            )

                    _output[_attribute.name] = attrib
                for outgoing_connection in obj.outgoing_connections:  # type: Connector
                    # print(
                    #     "Adding connection ", outgoing_connection.name, " to ", obj.name
                    # )
                    attrib = dict(
                        attribute_name=outgoing_connection.name,
                        attribute_type=document.get_object(
                            outgoing_connection.end_object_id
                        ).name,
                        attribute_cardinality=outgoing_connection.dest_card,
                        attribute_note=None,
                    )
                    _attr_ref = _ref.get_attribute(outgoing_connection.name)

                    if _attr_ref:
                        attrib["definition"] = _attr_ref.definition
                        attrib["c_code"] = _attr_ref.nci_c_code
                        attrib["pref_term"] = _attr_ref.preferred_term
                        if _attr_ref.has_value_list:
                            if _attr_ref.external_code_list:
                                attrib["external_value_list"] = _attr_ref.external_code_list
                            if _attr_ref.codelist_code:
                                attrib["codelist"] = _attr_ref.codelist_code
                    else:
                        print(
                            f"Reference not found for connection {outgoing_connection.name} in {_ref.entity_name}"
                        )
                    _output[attrib.get("attribute_name")] = attrib
                # connections = [
                #     connectors[cid]
                #     for cid in connectors
                #     if connectors[cid].start_object_id == _object_id
                # ]
                # for connection in connections:
                #     if connection.connector_type not in ("Association", "Generalization"):
                #         continue
                #     attrib = _output.setdefault(connection.name, {})
                #     if not attrib:
                #         attrib = dict(
                #             attribute_name=connection.name,
                #             attribute_type=objects[connection.end_object_id].name,
                #             attribute_cardinality=connection.source_card,
                #             note=None,
                #         )
                #     _output[connection.name] = attrib
                for _, attrib in enumerate(_output.values()):
                    _codelist = attrib.get("codelist")
                    if _codelist:
                        _cl = ct_content.get(_codelist)
                        if _codelist != "CNEW":
                            _codelist_value = '=HYPERLINK("#{}!A2","{}")'.format(
                                _codelist, _codelist
                            )
                        else:
                            _codelist_value = "CNEW"
                    else:
                        _codelist_value = ""
                    write_cell(sheet, row=row_num, column=1, value=obj.name)
                    write_cell(
                        sheet, row=row_num, column=2, value=attrib["attribute_name"]
                    )
                    write_cell(
                        sheet, row=row_num, column=3, value=attrib["attribute_type"]
                    )
                    write_cell(
                        sheet,
                        row=row_num,
                        column=4,
                        value=attrib["attribute_cardinality"],
                    )
                    if attrib["attribute_note"]:
                        write_cell(
                            sheet, row=row_num, column=5, value=attrib["attribute_note"]
                        )
                    write_cell(
                        sheet, row=row_num, column=6, value=attrib.get("c_code", "")
                    )
                    write_cell(
                        sheet, row=row_num, column=7, value=attrib.get("pref_term", "")
                    )
                    write_cell(
                        sheet, row=row_num, column=8, value=attrib.get("definition", "")
                    )
                    write_cell(sheet, row=row_num, column=9, value=_codelist_value)
                    write_cell(
                        sheet,
                        row=row_num,
                        column=10,
                        value=attrib.get("external_value_list", ""),
                    )
                    row_num += 1
    for c_code, _codelist in codelists.items():
        if c_code.strip().upper() == "CNEW":
            continue
        _sheet = doc.create_sheet(c_code)
        for idx, colheader in enumerate(
            ["Code", "Preferred Term", "Synonyms", "Definition"], start=1
        ):
            write_cell(_sheet, row=1, column=idx, value=colheader, header=True)
        for idx, item in enumerate(_codelist.items, start=2):
            # type item: PermissibleValue
            write_cell(_sheet, row=idx, column=1, value=item.concept_c_code)
            write_cell(_sheet, row=idx, column=2, value=item.preferred_term)
            write_cell(
                _sheet,
                row=idx,
                column=3,
                value=";".join(item.synonyms) if item.synonyms else "",
            )
            write_cell(_sheet, row=idx, column=4, value=item.definition)
    # Span the column widths
    for wks in doc.worksheets:
        _columns = cols.get(wks.title, {})
        for colnum, length in _columns.items():
            column_width = min([length, 50])
            wks.column_dimensions[get_column_letter(colnum)].width = column_width

    # create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    _name, _ = os.path.splitext(name)
    fname = os.path.join(output_dir, f"{_name}.xlsx")
    doc.save(fname)
    print(f"Generated USDM Excel file: {fname}")
